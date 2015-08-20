import math
import openpyxl
from openpyxl import *
import numpy as np
from numpy import *


# loads workbook. *MAKE THIS USER DEFINED
wb = load_workbook("/Users/caradonna/Downloads/paddle_locations.xlsx", data_only = True)


sheet_names = wb.get_sheet_names()

open("muoncounters.txt",'w').close()
muons = open("muoncounters.txt",'a')
detector_number = 0
x_z_flip = True

for q in range(0,len(sheet_names)):
    
    ws = wb[sheet_names[q]]
    #print sheet_names[q]
    muons.write("# " + sheet_names[q] + "\n")

    for a in range(1,100):
        for j in range(1,100):
            k = ws.cell(row = a, column = j).value
            if isinstance(k, basestring):
                if k.lower() == "x":
                    start_row = a + 1
                    start_column_tmp = j-1
                    end_column_tmp = j+1


                    for b in range(start_row,100):
                        if ws.cell(row = b, column = start_column_tmp).value == None:
                            end_row = b -1
                            break
                    break


    start_column = chr(start_column_tmp + ord('A'))
    end_column = chr(end_column_tmp +ord('A'))


    wsr = np.array(ws.range('%s%s:%s%s' %(start_column,start_row,end_column, end_row)))
    
    init_coord = np.zeros((end_row - start_row + 1,3))

    for i in range(0,3):
        for j in range(0,end_row - start_row + 1):
            init_coord[j][i] = wsr[j][i].value
            

    
        
    # declares an empty array that will be used to print out the coorinates of each corner of a given detector

    wall_or = 1

    detector_points = np.zeros((4,3))

    #due to how the points are given, some arrays need to be flipped vertically in order to be properly calculated
    if (init_coord[1][0] - init_coord[0][0] < 0.) or (init_coord[1][2] - init_coord[0][2] < 0):
        init_coord = np.flipud(init_coord)
        wall_or = -1

    #print init_coord
    #print init_coord[1][2] - init_coord[0][2]
    #print init_coord[1][0] - init_coord[0][0]
    
    # Fills array and then prints it
    # *possibly make it so that the same array is not reused
        
    if (math.fabs(init_coord[1][0] - init_coord[0][0]) == 14) and (init_coord[1][2] - init_coord[0][2] == 0):
        for i in range(0,end_row-start_row + 1):
            detector_points[0][0] = init_coord[i][0]
            detector_points[0][1] = init_coord[i][1]
            detector_points[0][2] = init_coord[i][2]


            detector_points[1][0] = init_coord[i][0] + ((12.82-10.65)/2)*wall_or
            detector_points[1][1] = init_coord[i][1] - (24.90)
            detector_points[1][2] = init_coord[i][2]
    
            detector_points[2][0] = init_coord[i][0] + (((12.82-10.65)/2) +10.65) *wall_or
            detector_points[2][1] = init_coord[i][1] -(24.90)
            detector_points[2][2] = init_coord[i][2]

            detector_points[3][0] = init_coord[i][0] + (12.82 * wall_or)
            detector_points[3][1] = init_coord[i][1]
            detector_points[3][2] = init_coord[i][2]

            
            if x_z_flip == True:
                for a in range(0,4):
                    detector_points[a][0] = detector_points[a][0] * -1
                    detector_points[a][2] = detector_points[a][2] * -1
        
            detector_points = detector_points *2.54
            
            muons.write(str(detector_number) + " 1 ")
            for j in range(0,4):
                for k in range(0,3):
                    muons.write(str(detector_points[j][k]) + " ")
            muons.write("\n")        
            detector_number += 1  
    
    if (math.fabs(init_coord[1][0] - init_coord[0][0]) == 6.75) and (init_coord[1][2] - init_coord[0][2] == 0):
        for i in range(0,end_row - start_row + 1):
    
            detector_points[0][0] = init_coord[i][0]
            detector_points[0][1] = init_coord[i][1]
            detector_points[0][2] = init_coord[i][2]

            detector_points[1][0] = init_coord[i][0]
            detector_points[1][1] = init_coord[i][1]
            detector_points[1][2] = init_coord[i][2] - 64.75

            detector_points[2][0] = init_coord[i][0] - 6.625
            detector_points[2][1] = init_coord[i][1]
            detector_points[2][2] = init_coord[i][2] - 64.75

            detector_points[3][0] = init_coord[i][0] - 6.625
            detector_points[3][1] = init_coord[i][1]    
            detector_points[3][2] = init_coord[i][2]
            
            if x_z_flip == True:
                for a in range(0,4):
                    detector_points[a][0] = detector_points[a][0] * -1
                    detector_points[a][2] = detector_points[a][2] * -1

            detector_points = detector_points *2.54
            
            muons.write(str(detector_number) + " 1 ")
            for j in range(0,4):
                for k in range(0,3):
                    muons.write(str(detector_points[j][k]) + " ")
            muons.write("\n")        
            detector_number += 1
            
    #printmath.fabs(init_coord[1][2] - init_coord[0][2])
    #print init_coord [1][0] - init_coord[0][0]
    
    #if (math.fabs(init_coord[1][2] - init_coord[0][2]) == 14.) and ((init_coord [1][0] - init_coord[0][0]) == 0.):
    if sheet_names[q].lower() == ("south wall - table 1") or sheet_names[q].lower() == ("north wall - table 1"):
        #print end_row - start_row + 1
        #print "test"
        for i in range(0,end_row - start_row + 1):
            detector_points[0][0] = init_coord[i][0]
            detector_points[0][1] = init_coord[i][1]
            detector_points[0][2] = init_coord[i][2]
        
            detector_points[1][0] = init_coord[i][0]
            detector_points[1][1] = init_coord[i][1] - 24.90
            detector_points[1][2] = init_coord[i][2] + ((12.82 - 10.65)/2) * wall_or

            detector_points[2][0] = init_coord[i][0]
            detector_points[2][1] = init_coord[i][1] - 24.90
            detector_points[2][2] = init_coord[i][2] + (((12.82 - 10.65)/2) + 10.65) * wall_or

            detector_points[3][0] = init_coord[i][0]
            detector_points[3][1] = init_coord[i][1]
            detector_points[3][2] = init_coord[i][2] + 12.82 * wall_or
            
            if x_z_flip == True:
                for a in range(0,4):
                    detector_points[a][0] = detector_points[a][0] * -1
                    detector_points[a][2] = detector_points[a][2] * -1

            detector_points = detector_points *2.54

            muons.write(str(detector_number) + " 1 ")
            for j in range(0,4):
                for k in range(0,3):
                    muons.write(str(detector_points[j][k]) + " ")
            muons.write("\n")        
            detector_number += 1  
    

    if (math.fabs(init_coord[1][2] - init_coord[0][2]) == 6.75) and (init_coord[1][0] - init_coord[0][0] == 0):
        for i in range(0,end_row-start_row + 1):
            detector_points[0][0] = init_coord[i][0] 
            detector_points[0][1] = init_coord[i][1]
            detector_points[0][2] = init_coord[i][2]

            detector_points[1][0] = init_coord[i][0] + 64.75
            detector_points[1][1] = init_coord[i][1]
            detector_points[1][2] = init_coord[i][2]

            detector_points[2][0] = init_coord[i][0] + 64.75
            detector_points[2][1] = init_coord[i][1]
            detector_points[2][2] = init_coord[i][2] - 6.625

            detector_points[3][0] = init_coord[i][0]
            detector_points[3][1] = init_coord[i][1]
            detector_points[3][2] = init_coord[i][2] - 6.625
            
            if x_z_flip == True:
                for a in range(0,4):
                    detector_points[a][0] = detector_points[a][0] * -1
                    detector_points[a][2] = detector_points[a][2] * -1

            detector_points = detector_points *2.54

            muons.write(str(detector_number) + " 1 ")
            for j in range(0,4):
                for k in range(0,3):
                    muons.write(str(detector_points[j][k]) + " ")
            muons.write("\n")        
            detector_number += 1        

        

    
